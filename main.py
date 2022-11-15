import sqlite3
import openpyxl
from prettytable import from_db_cursor

class ExcelToSqlite(object):
    """Общий класс работы для преобразования файла excel в базу данных SQLite3 """

    def __init__(self, dbName):
        print("Инициализация экземпляра базы данных SQLite3")
        super(ExcelToSqlite, self).__init__()
        print("     Установка соединения с базой данных\n")
        self.conn = sqlite3.connect(dbName)
        self.cursor = self.conn.cursor()

    def __del__(self):
        print("Завершение сеанса соединения с базой данных. Работа завершена. ")
        self.cursor.close()
        self.conn.close()

    def ExcelToDb(self, tableName, excelName):
        """
        Метод для преобразования Excel в таблицу базы данных sqlite
        :param tableName: имя таблицы базы данных
        """
        print("Преобразования файла excel в базу данных:")
        self.tableName = tableName
        self.__CreateTable(tableName)
        print("     2)Считывание строк таблицы excel и наполнение базы данных...")
        file_to_read = openpyxl.load_workbook(excelName, data_only=True)
        sheet = file_to_read['Лист1']
        # Цикл по строкам начиная со второй (в первой заголовки)
        for row in range(4, sheet.max_row + 1):
            # Объявление списка
            data = []
            # Цикл по столбцам от 1 до 10 (11 не включая)
            for col in range(1, 11):
                # value содержит значение ячейки с координатами row col
                value = sheet.cell(row, col).value
                # Список который мы потом будем добавлять
                data.append(value)
            self.__Insert(data)
        print("     3)База данных наполнена")

    def __CreateTable(self, tableName):
        """
        Метод для создания таблицы в базе данных
        :param tableName: имя таблицы
        :return:
        """
        print("     1)Создание таблицы " + tableName)

        sql = 'CREATE TABLE IF NOT EXISTS %s (date_id int, company str, TotalfactQLiq int, TotalfactQOil int, TotalfactQLiqQOil int, TotalforecastQLiq int, TotalforecastQOil int, TotalforecastQLiqQOil int)' % (self.tableName)
        self.cursor.execute(sql)
        self.conn.commit()

    def __Insert(self, data):
        """
        Метод для наполнения базы данных
        :param data: список значений
        """
        self.cursor.execute(f"INSERT INTO {tableName} VALUES (?, ?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2]+data[3], data[4]+data[5],data[2]+data[3]+data[4]+data[5], data[6]+data[7], data[8]+data[9], data[6]+data[7]+data[8]+data[9]))
        self.conn.commit()

    def Query(self, tableName):
        """
        Данные запроса в таблицах базы данных
        :param tableName: имя таблицы
        """
        self.cursor.execute('select * from %s' % (tableName))
        mytable = from_db_cursor(self.cursor)
        mytable.field_names = ["Номер дня",
                               "Название компании",
                               "Объем жидкости(факт)",
                               "Объем нефти(факт)",
                               "Общая жидкости и нефти(факт)",
                               "Объем жидкости(прогноз)",
                               "Объем нефти(прогноз)",
                               "Общая жидкости и нефти(прогноз)"
                               ]
        print(mytable)


sqlite3.connect("dbName.db")    # Создайте базу данных самостоятельно с помощью sqlite3.
dbName = "dbName"               # Установите имя базы данных в качестве переменной функции.
tableName = "testName"          # Установите имя таблицы базы данных.
excelName = "Задание бек.xlsx"  # Установите имя excel файла.

test = ExcelToSqlite(dbName)
test.ExcelToDb(tableName, excelName)
test.Query(tableName)
